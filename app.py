import os
import glob
import hashlib
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()  # Must run before any local imports that read env vars

from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, flash, send_file)
from concurrent.futures import ThreadPoolExecutor, as_completed

from database import (init_db, save_report, get_report, get_report_age,
                      get_all_reports, delete_report, save_market_items,
                      get_market_items, get_profile, save_profile,
                      save_youtube_cache, get_youtube_cache,
                      save_calculator_data, get_calculator_data,
                      save_comparison, get_all_comparisons,
                      update_comparison, delete_comparison)
from scraper import scrape_city_data, resolve_city_slug, suggest_cities, resolve_true_slug
from youtube_api import get_city_videos


def _norm_name(s):
    return ' '.join(s.replace('²', '2').split()).lower()


def _rank_color(rank, max_rank):
    """Dark green (rank 1) → amber → dark red (last rank) gradient."""
    if max_rank <= 1:
        return '#1b5e20'
    t = (rank - 1) / (max_rank - 1)
    if t <= 0.5:
        t2 = t * 2
        r = int(27 + t2 * (245 - 27))
        g = int(94 + t2 * (127 - 94))
        b = int(32 + t2 * (23 - 32))
    else:
        t2 = (t - 0.5) * 2
        r = int(245 + t2 * (183 - 245))
        g = int(127 + t2 * (28 - 127))
        b = int(23 + t2 * (28 - 23))
    return f'#{r:02x}{g:02x}{b:02x}'


def _freshness(created_at_str):
    """3-tier freshness: 0-30 Fresh, 31-90 Consider refresh, >90 Old."""
    try:
        created = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
        days = (datetime.now() - created).days
        date_str = created.strftime('%d/%m/%Y')
    except Exception:
        days = 0
        date_str = ''

    if days <= 30:
        return {'level': 'fresh', 'cls': 'success', 'days': days,
                'label': 'Fresh data', 'date_str': date_str}
    elif days <= 90:
        return {'level': 'caution', 'cls': 'warning', 'days': days,
                'label': 'Consider refresh', 'date_str': date_str}
    else:
        return {'level': 'danger', 'cls': 'danger', 'days': days,
                'label': 'Old data - refresh recommended', 'date_str': date_str}

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'soft-landing-dev-key-change-me')

CURRENCIES = [
    ('USD', 'USD — US Dollar'),
    ('EUR', 'EUR — Euro'),
    ('GBP', 'GBP — British Pound'),
    ('JPY', 'JPY — Japanese Yen'),
    ('BRL', 'BRL — Brazilian Real'),
    ('CAD', 'CAD — Canadian Dollar'),
    ('AUD', 'AUD — Australian Dollar'),
    ('CHF', 'CHF — Swiss Franc'),
    ('CNY', 'CNY — Chinese Yuan'),
    ('INR', 'INR — Indian Rupee'),
    ('MXN', 'MXN — Mexican Peso'),
    ('SGD', 'SGD — Singapore Dollar'),
    ('HKD', 'HKD — Hong Kong Dollar'),
    ('NZD', 'NZD — New Zealand Dollar'),
    ('SEK', 'SEK — Swedish Krona'),
    ('NOK', 'NOK — Norwegian Krone'),
    ('DKK', 'DKK — Danish Krone'),
    ('THB', 'THB — Thai Baht'),
    ('MYR', 'MYR — Malaysian Ringgit'),
    ('PHP', 'PHP — Philippine Peso'),
    ('KRW', 'KRW — South Korean Won'),
    ('ZAR', 'ZAR — South African Rand'),
    ('AED', 'AED — UAE Dirham'),
    ('TRY', 'TRY — Turkish Lira'),
    ('PLN', 'PLN — Polish Zloty'),
    ('CZK', 'CZK — Czech Koruna'),
    ('ARS', 'ARS — Argentine Peso'),
    ('CLP', 'CLP — Chilean Peso'),
    ('COP', 'COP — Colombian Peso'),
    ('PEN', 'PEN — Peruvian Sol'),
    ('VND', 'VND — Vietnamese Dong'),
    ('IDR', 'IDR — Indonesian Rupiah'),
]

init_db()

# ── HTML report file cache ─────────────────────────────────────────────────────
_CACHE_DIR = os.environ.get('CACHE_DIR', os.path.join(os.path.dirname(__file__), 'city_reports'))
os.makedirs(_CACHE_DIR, exist_ok=True)


def _profile_hash(currency, budget):
    return hashlib.md5(f'{currency}:{budget:.2f}'.encode()).hexdigest()[:8]


def _get_html_cache(city_slug, phash):
    path = os.path.join(_CACHE_DIR, f'{city_slug}_{phash}.html')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return f.read()
    return None


def _save_html_cache(city_slug, phash, html):
    for old in glob.glob(os.path.join(_CACHE_DIR, f'{city_slug}_*.html')):
        try:
            os.remove(old)
        except OSError:
            pass
    try:
        with open(os.path.join(_CACHE_DIR, f'{city_slug}_{phash}.html'), 'w', encoding='utf-8') as f:
            f.write(html)
    except OSError:
        pass


def _clear_html_cache(city_slug):
    for old in glob.glob(os.path.join(_CACHE_DIR, f'{city_slug}_*.html')):
        try:
            os.remove(old)
        except OSError:
            pass


# ── Country name → Unicode flag emoji ─────────────────────────────────────────
_COUNTRY_ISO = {
    'Afghanistan': 'AF', 'Albania': 'AL', 'Algeria': 'DZ', 'Angola': 'AO',
    'Argentina': 'AR', 'Armenia': 'AM', 'Australia': 'AU', 'Austria': 'AT',
    'Azerbaijan': 'AZ', 'Bahrain': 'BH', 'Bangladesh': 'BD', 'Belarus': 'BY',
    'Belgium': 'BE', 'Bolivia': 'BO', 'Bosnia And Herzegovina': 'BA',
    'Brazil': 'BR', 'Bulgaria': 'BG', 'Cambodia': 'KH', 'Cameroon': 'CM',
    'Canada': 'CA', 'Chile': 'CL', 'China': 'CN', 'Colombia': 'CO',
    'Costa Rica': 'CR', 'Croatia': 'HR', 'Cuba': 'CU', 'Cyprus': 'CY',
    'Czech Republic': 'CZ', 'Denmark': 'DK', 'Dominican Republic': 'DO',
    'Ecuador': 'EC', 'Egypt': 'EG', 'El Salvador': 'SV', 'Estonia': 'EE',
    'Ethiopia': 'ET', 'Finland': 'FI', 'France': 'FR', 'Georgia': 'GE',
    'Germany': 'DE', 'Ghana': 'GH', 'Greece': 'GR', 'Guatemala': 'GT',
    'Honduras': 'HN', 'Hong Kong': 'HK', 'Hungary': 'HU', 'Iceland': 'IS',
    'India': 'IN', 'Indonesia': 'ID', 'Iran': 'IR', 'Iraq': 'IQ',
    'Ireland': 'IE', 'Israel': 'IL', 'Italy': 'IT', 'Jamaica': 'JM',
    'Japan': 'JP', 'Jordan': 'JO', 'Kazakhstan': 'KZ', 'Kenya': 'KE',
    'Kuwait': 'KW', 'Latvia': 'LV', 'Lebanon': 'LB', 'Lithuania': 'LT',
    'Luxembourg': 'LU', 'Malaysia': 'MY', 'Malta': 'MT', 'Mexico': 'MX',
    'Moldova': 'MD', 'Mongolia': 'MN', 'Montenegro': 'ME', 'Morocco': 'MA',
    'Mozambique': 'MZ', 'Myanmar': 'MM', 'Nepal': 'NP', 'Netherlands': 'NL',
    'New Zealand': 'NZ', 'Nicaragua': 'NI', 'Nigeria': 'NG', 'Norway': 'NO',
    'Oman': 'OM', 'Pakistan': 'PK', 'Panama': 'PA', 'Paraguay': 'PY',
    'Peru': 'PE', 'Philippines': 'PH', 'Poland': 'PL', 'Portugal': 'PT',
    'Qatar': 'QA', 'Romania': 'RO', 'Russia': 'RU', 'Saudi Arabia': 'SA',
    'Senegal': 'SN', 'Serbia': 'RS', 'Singapore': 'SG', 'Slovakia': 'SK',
    'Slovenia': 'SI', 'South Africa': 'ZA', 'South Korea': 'KR', 'Spain': 'ES',
    'Sri Lanka': 'LK', 'Sudan': 'SD', 'Sweden': 'SE', 'Switzerland': 'CH',
    'Taiwan': 'TW', 'Tanzania': 'TZ', 'Thailand': 'TH', 'Tunisia': 'TN',
    'Turkey': 'TR', 'Uganda': 'UG', 'Ukraine': 'UA',
    'United Arab Emirates': 'AE', 'United Kingdom': 'GB',
    'United States': 'US', 'Uruguay': 'UY', 'Uzbekistan': 'UZ',
    'Venezuela': 'VE', 'Vietnam': 'VN', 'Yemen': 'YE', 'Zambia': 'ZM',
    'Zimbabwe': 'ZW',
}


_COUNTRY_CONTINENT = {
    # Americas
    'Argentina': 'Americas', 'Bolivia': 'Americas', 'Brazil': 'Americas',
    'Canada': 'Americas', 'Chile': 'Americas', 'Colombia': 'Americas',
    'Costa Rica': 'Americas', 'Cuba': 'Americas', 'Dominican Republic': 'Americas',
    'Ecuador': 'Americas', 'El Salvador': 'Americas', 'Guatemala': 'Americas',
    'Honduras': 'Americas', 'Jamaica': 'Americas', 'Mexico': 'Americas',
    'Nicaragua': 'Americas', 'Panama': 'Americas', 'Paraguay': 'Americas',
    'Peru': 'Americas', 'United States': 'Americas', 'Uruguay': 'Americas',
    'Venezuela': 'Americas',
    # Europe
    'Albania': 'Europe', 'Austria': 'Europe', 'Belarus': 'Europe',
    'Belgium': 'Europe', 'Bosnia And Herzegovina': 'Europe', 'Bulgaria': 'Europe',
    'Croatia': 'Europe', 'Cyprus': 'Europe', 'Czech Republic': 'Europe',
    'Denmark': 'Europe', 'Estonia': 'Europe', 'Finland': 'Europe',
    'France': 'Europe', 'Germany': 'Europe', 'Greece': 'Europe',
    'Hungary': 'Europe', 'Iceland': 'Europe', 'Ireland': 'Europe',
    'Italy': 'Europe', 'Latvia': 'Europe', 'Lithuania': 'Europe',
    'Luxembourg': 'Europe', 'Malta': 'Europe', 'Moldova': 'Europe',
    'Montenegro': 'Europe', 'Netherlands': 'Europe', 'Norway': 'Europe',
    'Poland': 'Europe', 'Portugal': 'Europe', 'Romania': 'Europe',
    'Russia': 'Europe', 'Serbia': 'Europe', 'Slovakia': 'Europe',
    'Slovenia': 'Europe', 'Spain': 'Europe', 'Sweden': 'Europe',
    'Switzerland': 'Europe', 'Ukraine': 'Europe', 'United Kingdom': 'Europe',
    # Asia
    'Afghanistan': 'Asia', 'Armenia': 'Asia', 'Azerbaijan': 'Asia',
    'Bahrain': 'Asia', 'Bangladesh': 'Asia', 'Cambodia': 'Asia',
    'China': 'Asia', 'Georgia': 'Asia', 'Hong Kong': 'Asia',
    'India': 'Asia', 'Indonesia': 'Asia', 'Iran': 'Asia',
    'Iraq': 'Asia', 'Israel': 'Asia', 'Japan': 'Asia',
    'Jordan': 'Asia', 'Kazakhstan': 'Asia', 'Kuwait': 'Asia',
    'Malaysia': 'Asia', 'Mongolia': 'Asia', 'Myanmar': 'Asia',
    'Nepal': 'Asia', 'Oman': 'Asia', 'Pakistan': 'Asia',
    'Philippines': 'Asia', 'Qatar': 'Asia', 'Saudi Arabia': 'Asia',
    'Singapore': 'Asia', 'South Korea': 'Asia', 'Sri Lanka': 'Asia',
    'Taiwan': 'Asia', 'Thailand': 'Asia', 'Turkey': 'Asia',
    'United Arab Emirates': 'Asia', 'Uzbekistan': 'Asia', 'Vietnam': 'Asia',
    'Yemen': 'Asia',
    # Africa
    'Algeria': 'Africa', 'Angola': 'Africa', 'Cameroon': 'Africa',
    'Egypt': 'Africa', 'Ethiopia': 'Africa', 'Ghana': 'Africa',
    'Kenya': 'Africa', 'Morocco': 'Africa', 'Mozambique': 'Africa',
    'Nigeria': 'Africa', 'Senegal': 'Africa', 'South Africa': 'Africa',
    'Sudan': 'Africa', 'Tanzania': 'Africa', 'Tunisia': 'Africa',
    'Uganda': 'Africa', 'Zambia': 'Africa', 'Zimbabwe': 'Africa',
    # Oceania
    'Australia': 'Oceania', 'New Zealand': 'Oceania',
}

_CONTINENT_ORDER = ['Americas', 'Europe', 'Asia', 'Africa', 'Oceania', 'Other']


def _country_flag(country_name):
    code = _COUNTRY_ISO.get(country_name, '')
    if not code:
        return ''
    return ''.join(chr(0x1F1E6 + ord(c) - ord('A')) for c in code)


app.jinja_env.filters['flag'] = _country_flag


@app.route('/')
def index():
    recent = get_all_reports(limit=6)
    return render_template('index.html', recent_reports=recent)


@app.route('/search', methods=['POST'])
def search():
    city = request.form.get('city', '').strip()
    if not city:
        flash('Please enter a city name.', 'warning')
        return redirect(url_for('index'))
    city_id = request.form.get('city_id', '').strip()
    # Best path: use Numbeo's dispatcher with city_id to get the exact page slug
    if city_id:
        slug = (resolve_true_slug(city, city_id)
                or request.form.get('resolved_slug', '').strip()
                or city.replace(' ', '-'))
    else:
        # Manual input fallback: resolve via autocomplete, then simple slugify
        slug = resolve_city_slug(city) or city.replace(' ', '-')
    return redirect(url_for('report', city=slug))


@app.route('/api/city-suggest')
def api_city_suggest():
    term = request.args.get('term', '').strip()
    if len(term) < 2:
        return jsonify([])
    return jsonify(suggest_cities(term))


@app.route('/report/<path:city>')
def report(city):
    city_display = city.replace('-', ' ').title()
    refresh = request.args.get('refresh') == '1'
    profile = get_profile()
    profile_currency = profile.get('currency', 'USD')
    profile_budget = float(profile.get('budget', 0) or 0)
    p_hash = _profile_hash(profile_currency, profile_budget)

    if refresh:
        _clear_html_cache(city)

    # ── Step 1: serve pre-rendered HTML file if it exists (fastest path) ─────
    if not refresh:
        cached_html = _get_html_cache(city, p_hash)
        if cached_html:
            return cached_html

    # ── Step 2: no HTML file — check SQLite for city data ────────────────────
    data = None
    from_cache = False
    if not refresh:
        cached = get_report(city_display)
        if cached and cached.get('currency') == profile_currency and 'country' in cached:
            data = cached
            from_cache = True
            created_at = get_report_age(city_display) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── Step 3: not in SQLite (or refresh) — scrape Numbeo ───────────────────
    if data is None:
        data = scrape_city_data(city, currency=profile_currency)
        if data:
            save_report(city_display, data)
            created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            flash(
                f'Could not retrieve data for "{city_display}". '
                'Check the spelling or try another city.',
                'danger'
            )
            return redirect(url_for('index'))

    # ── YouTube videos ────────────────────────────────────────────────────────
    def _has_videos(v):
        return bool(v) and any(lst for lst in v.values())

    if refresh:
        videos = get_city_videos(city_display)
        if _has_videos(videos):
            save_youtube_cache(city_display, videos)
    else:
        videos = get_youtube_cache(city_display)
        if not _has_videos(videos):
            videos = get_city_videos(city_display)
            if _has_videos(videos):
                save_youtube_cache(city_display, videos)

    # ── Render, cache to HTML file, return ────────────────────────────────────
    rendered = render_template(
        'report.html',
        city=city_display,
        city_slug=city,
        data=data,
        videos=videos,
        from_cache=from_cache,
        profile_budget=profile_budget,
        freshness=_freshness(created_at),
        cached_at=created_at,
        calculator_data=get_calculator_data(),
    )
    _save_html_cache(city, p_hash, rendered)
    return rendered


@app.route('/history')
def history():
    reports = get_all_reports()
    return render_template('history.html', reports=reports)


@app.route('/compare')
def compare():
    cities = request.args.getlist('city')
    profile_data = get_profile()
    profile_currency = profile_data.get('currency', 'USD')
    budget_total = float(profile_data.get('budget', 0) or 0)
    def _fetch(c):
        c_display = c.replace('-', ' ').title()
        d = get_report(c_display)
        if d and d.get('currency') != profile_currency:
            d = None
        if not d:
            d = scrape_city_data(c, currency=profile_currency)
        return c_display, d

    comparison = {}
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(_fetch, c): c for c in cities}
        for future in as_completed(futures):
            c_display, d = future.result()
            if d:
                save_report(c_display, d)
                comparison[c_display] = d

    # Compute affordability rankings: count how many items each city prices lowest
    affordability_rank = {}
    least_affordable_rank = 0
    if len(comparison) >= 2:
        city_cheapest = {city: 0 for city in comparison}
        all_item_prices = {}
        for city, city_data in comparison.items():
            for cat, items in city_data.get('categories', {}).items():
                for item in items:
                    if not item.get('price'):
                        continue
                    key = (cat, item['name'])
                    if key not in all_item_prices:
                        all_item_prices[key] = {}
                    all_item_prices[key][city] = item['price']
        for city_prices in all_item_prices.values():
            if len(city_prices) < 2:
                continue
            min_price = min(city_prices.values())
            for city, price in city_prices.items():
                if price == min_price:
                    city_cheapest[city] += 1
        sorted_counts = sorted(city_cheapest.values(), reverse=True)
        affordability_rank = {
            city: sorted_counts.index(count) + 1
            for city, count in city_cheapest.items()
        }
        least_affordable_rank = max(affordability_rank.values()) if affordability_rank else 0

    # Summary insights
    most_affordable_city = next(
        (c for c, r in affordability_rank.items() if r == 1),
        next(iter(comparison), '')
    )
    cheapest_restaurants_city = ''
    cheapest_markets_city = ''
    city_salaries = {}
    if comparison:
        rest_avgs, mkt_avgs = {}, {}
        for city, d in comparison.items():
            for cat_name, items in d.get('categories', {}).items():
                prices = [i['price'] for i in items if i.get('price')]
                if prices:
                    avg = sum(prices) / len(prices)
                    if cat_name == 'Restaurants':
                        rest_avgs[city] = avg
                    elif cat_name == 'Markets':
                        mkt_avgs[city] = avg
            for item in d.get('categories', {}).get('Salaries And Financing', []):
                if 'Average Monthly Net Salary' in item.get('name', ''):
                    city_salaries[city] = item.get('price', 0)
                    break
        if rest_avgs:
            cheapest_restaurants_city = min(rest_avgs, key=rest_avgs.get)
        if mkt_avgs:
            cheapest_markets_city = min(mkt_avgs, key=mkt_avgs.get)

    # Rank cities by budget/salary ratio (rank 1 = budget goes furthest vs local salary)
    salary_rank = {}
    salary_rank_max = 0
    if budget_total > 0 and city_salaries:
        ratios = {city: budget_total / sal for city, sal in city_salaries.items() if sal > 0}
        sorted_by_ratio = sorted(ratios, key=lambda c: ratios[c], reverse=True)
        salary_rank = {city: i + 1 for i, city in enumerate(sorted_by_ratio)}
        salary_rank_max = len(salary_rank)

    # Cost projection rankings — mirrors JS calculator logic
    calc_data_for_rank = get_calculator_data()
    calc_projection_rank = {}
    calc_projection_rank_max = 0
    if calc_data_for_rank and comparison:
        city_price_lookup = {}
        for city, city_data in comparison.items():
            city_price_lookup[city] = {}
            for cat, items in city_data.get('categories', {}).items():
                for item in items:
                    if item.get('price'):
                        key = _norm_name(item['name'])
                        if key not in city_price_lookup[city]:
                            city_price_lookup[city][key] = item['price']
        city_totals = {city: 0.0 for city in comparison}
        for calc_name, qty in calc_data_for_rank.items():
            if not qty:
                continue
            key = _norm_name(calc_name)
            for city in comparison:
                price = city_price_lookup[city].get(key)
                if price:
                    city_totals[city] += price * qty
        ranked = sorted(
            [(c, t) for c, t in city_totals.items() if t > 0],
            key=lambda x: x[1]
        )
        if ranked:
            calc_projection_rank = {city: i + 1 for i, (city, _) in enumerate(ranked)}
            calc_projection_rank_max = len(calc_projection_rank)

    all_reports = get_all_reports()

    # Group reports by continent → country for the city selector
    grouped_reports = {}
    for r in all_reports:
        cont = _COUNTRY_CONTINENT.get(r.get('country', ''), 'Other')
        country = r.get('country') or 'Unknown'
        grouped_reports.setdefault(cont, {}).setdefault(country, []).append(r)

    return render_template('compare.html',
                           comparison=comparison,
                           all_reports=all_reports,
                           grouped_reports=grouped_reports,
                           continent_order=_CONTINENT_ORDER,
                           selected=cities,
                           affordability_rank=affordability_rank,
                           least_affordable_rank=least_affordable_rank,
                           most_affordable_city=most_affordable_city,
                           cheapest_restaurants_city=cheapest_restaurants_city,
                           cheapest_markets_city=cheapest_markets_city,
                           city_salaries=city_salaries,
                           budget_total=budget_total,
                           profile_currency=profile_currency,
                           salary_rank=salary_rank,
                           salary_rank_max=salary_rank_max,
                           calc_projection_rank=calc_projection_rank,
                           calc_projection_rank_max=calc_projection_rank_max,
                           calculator_data=get_calculator_data(),
                           saved_comparisons=get_all_comparisons())


@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        currency = request.form.get('currency', 'USD')
        budget = float(request.form.get('budget', 0) or 0)
        if not name or not email:
            flash('Name and email are required.', 'warning')
            return redirect(url_for('profile'))
        valid_codes = {c[0] for c in CURRENCIES}
        if currency not in valid_codes:
            currency = 'USD'
        save_profile(name, email, currency, budget)
        flash('Profile saved successfully.', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html',
                           profile=get_profile(),
                           currencies=CURRENCIES,
                           market_items=get_market_items(),
                           calculator_data=get_calculator_data())


# ── API endpoints ──────────────────────────────────────────────────────────────

@app.route('/api/market', methods=['POST'])
def api_market():
    items = request.json.get('items', [])
    save_market_items(items)
    return jsonify({'ok': True})


@app.route('/api/calculator', methods=['GET', 'POST'])
def api_calculator():
    if request.method == 'POST':
        body = request.json or {}
        data = body.get('data', {})
        if not isinstance(data, dict):
            return jsonify({'ok': False}), 400
        save_calculator_data(data)
        return jsonify({'ok': True})
    return jsonify(get_calculator_data())


@app.route('/api/report/<int:report_id>', methods=['DELETE'])
def api_delete(report_id):
    delete_report(report_id)
    return jsonify({'ok': True})


@app.route('/api/saved-comparisons', methods=['GET', 'POST'])
def api_saved_comparisons():
    if request.method == 'POST':
        body = request.json or {}
        name = body.get('name', '').strip()
        cities = body.get('cities', [])
        if not name or not cities:
            return jsonify({'ok': False}), 400
        new_id = save_comparison(name, cities)
        return jsonify({'ok': True, 'id': new_id})
    return jsonify(get_all_comparisons())


@app.route('/api/saved-comparisons/<int:comp_id>', methods=['PUT', 'DELETE'])
def api_saved_comparison(comp_id):
    if request.method == 'DELETE':
        delete_comparison(comp_id)
        return jsonify({'ok': True})
    body = request.json or {}
    name = body.get('name', '').strip()
    cities = body.get('cities', [])
    if not name:
        return jsonify({'ok': False}), 400
    update_comparison(comp_id, name, cities)
    return jsonify({'ok': True})


@app.route('/compare/download')
def compare_download():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cities = request.args.getlist('city')
    profile_data = get_profile()
    profile_currency = profile_data.get('currency', 'USD')

    def _fetch(c):
        c_display = c.replace('-', ' ').title()
        d = get_report(c_display)
        if d and d.get('currency') != profile_currency:
            d = None
        if not d:
            d = scrape_city_data(c, currency=profile_currency)
        return c_display, d

    comparison = {}
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(_fetch, c): c for c in cities}
        for future in as_completed(futures):
            c_display, d = future.result()
            if d:
                comparison[c_display] = d

    if not comparison:
        flash('No data available to download.', 'warning')
        return redirect(url_for('compare'))

    city_names = list(comparison.keys())
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1565C0')
    green_fill = PatternFill('solid', fgColor='C8E6C9')
    red_fill = PatternFill('solid', fgColor='FFCDD2')

    def _style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Summary sheet — average price per category per city
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    all_cats = []
    for d in comparison.values():
        for cat in d.get('categories', {}):
            if cat not in all_cats:
                all_cats.append(cat)

    ws_sum.cell(1, 1, 'Category')
    for col, city in enumerate(city_names, 2):
        ws_sum.cell(1, col, city)
    _style_header(ws_sum, 1, 1 + len(city_names))

    for row, cat in enumerate(all_cats, 2):
        ws_sum.cell(row, 1, cat)
        cat_prices = []
        for col, city in enumerate(city_names, 2):
            items = comparison[city].get('categories', {}).get(cat, [])
            prices = [i['price'] for i in items if i.get('price')]
            avg = round(sum(prices) / len(prices), 2) if prices else None
            cell = ws_sum.cell(row, col, avg)
            cell.alignment = Alignment(horizontal='center')
            if avg is not None:
                cat_prices.append((col, avg))
        if len(cat_prices) > 1:
            min_v = min(v for _, v in cat_prices)
            max_v = max(v for _, v in cat_prices)
            for col, v in cat_prices:
                ws_sum.cell(row, col).fill = green_fill if v == min_v else (red_fill if v == max_v else PatternFill())

    for col in range(1, 2 + len(city_names)):
        ws_sum.column_dimensions[get_column_letter(col)].width = 22

    # One sheet per category
    for cat in all_cats:
        safe_name = cat[:31].replace('/', '-')
        ws = wb.create_sheet(title=safe_name)
        ws.cell(1, 1, 'Item')
        for col, city in enumerate(city_names, 2):
            ws.cell(1, col, city)
        _style_header(ws, 1, 1 + len(city_names))

        all_items = []
        for city in city_names:
            for item in comparison[city].get('categories', {}).get(cat, []):
                if item['name'] not in all_items:
                    all_items.append(item['name'])

        for row, item_name in enumerate(all_items, 2):
            ws.cell(row, 1, item_name)
            row_prices = []
            for col, city in enumerate(city_names, 2):
                price = next(
                    (i['price'] for i in comparison[city].get('categories', {}).get(cat, [])
                     if i['name'] == item_name and i.get('price')),
                    None
                )
                cell = ws.cell(row, col, price)
                cell.alignment = Alignment(horizontal='center')
                if price is not None:
                    row_prices.append((col, price))
            if len(row_prices) > 1:
                min_v = min(v for _, v in row_prices)
                max_v = max(v for _, v in row_prices)
                for col, v in row_prices:
                    ws.cell(row, col).fill = green_fill if v == min_v else (red_fill if v == max_v else PatternFill())

        for col in range(1, 2 + len(city_names)):
            ws.column_dimensions[get_column_letter(col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = 'comparison_' + '_'.join(c.replace(' ', '-') for c in city_names[:3]) + '.xlsx'
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=os.environ.get('FLASK_ENV') != 'production',
            host='0.0.0.0', port=port)
